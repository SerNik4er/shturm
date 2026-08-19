import os
import asyncio
import logging
import tempfile
from datetime import datetime
from openpyxl import Workbook, load_workbook
from vkbottle.bot import Bot, Message
from vkbottle import Keyboard, KeyboardButtonColor, Text, Uploader
from dotenv import load_dotenv
load_dotenv()

# Отключаем логирование
logging.getLogger("vkbottle").setLevel(logging.CRITICAL)
logging.getLogger("loguru").setLevel(logging.CRITICAL)

# ТОКЕН ИЗ ПЕРЕМЕННОЙ ОКРУЖЕНИЯ
TOKEN = os.getenv("VK_BOT_TOKEN")

if not TOKEN:
    raise ValueError("❌ Токен не найден! Установите VK_BOT_TOKEN")

# ===== АДМИНИСТРАТОРЫ =====
ADMIN_IDS = [
    164876852,   # Админ 1
    21212595,# Админ 2 (ЗАМЕНИТЕ НА РЕАЛЬНЫЙ ID!)
    531011063,
]

# ===== СПИСОК ГРУПП ДЛЯ ПРОВЕРКИ ПОДПИСКИ =====
REQUIRED_GROUPS = [
    237204348,
    108117049,
]

# ===== СПИСОК МЕРОПРИЯТИЙ =====
EVENTS = [
    "Штурмовой бой - дети",
   
]

NOT_SUBSCRIBED_TEXT = (
    "❌ Вы не подписаны на обязательные группы!\n\n"
    "Чтобы продолжить, подпишитесь на все группы:\n"
)

bot = Bot(token=TOKEN)

DATA_FILE = "users_data.xlsx"
user_data_temp = {}

# ===== ФУНКЦИИ РАБОТЫ С EXCEL =====

def init_excel():
    """Создает Excel-файл с заголовками, если его нет"""
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        headers = ["ID", "ФИО", "Телефон", "Возраст", "Мероприятие", "Дата регистрации"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        wb.save(DATA_FILE)

def save_user_data(user_id, full_name, phone, age, event):
    """Сохраняет данные пользователя в Excel"""
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=user_id)
    ws.cell(row=row, column=2, value=full_name)
    ws.cell(row=row, column=3, value=phone)
    ws.cell(row=row, column=4, value=age)
    ws.cell(row=row, column=5, value=event)
    ws.cell(row=row, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    wb.save(DATA_FILE)

def is_user_registered(user_id: int) -> bool:
    """Проверяет, зарегистрирован ли пользователь"""
    if not os.path.exists(DATA_FILE):
        return False
    
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == user_id:
            return True
    return False

# ===== КЛАВИАТУРЫ =====

def get_main_keyboard():
    keyboard = Keyboard(inline=False)
    keyboard.add(Text("📝 Регистрация"), color=KeyboardButtonColor.PRIMARY)
    keyboard.row()
    keyboard.add(Text("📊 Посмотреть данные"), color=KeyboardButtonColor.SECONDARY)
    keyboard.row()
    keyboard.add(Text("❓ Помощь"), color=KeyboardButtonColor.SECONDARY)
    return keyboard

def get_cancel_keyboard():
    keyboard = Keyboard(inline=False)
    keyboard.add(Text("❌ Отмена"), color=KeyboardButtonColor.NEGATIVE)
    return keyboard

def get_subscription_keyboard():
    keyboard = Keyboard(inline=False)
    keyboard.add(Text("✅ Проверить подписку"), color=KeyboardButtonColor.POSITIVE)
    keyboard.row()
    keyboard.add(Text("❌ Отмена"), color=KeyboardButtonColor.NEGATIVE)
    return keyboard

def get_event_keyboard():
    keyboard = Keyboard(inline=False)
    for i, event in enumerate(EVENTS):
        keyboard.add(Text(event), color=KeyboardButtonColor.PRIMARY)
        if i % 2 == 1 and i < len(EVENTS) - 1:
            keyboard.row()
    if len(EVENTS) % 2 == 1:
        keyboard.row()
    keyboard.add(Text("❌ Отмена"), color=KeyboardButtonColor.NEGATIVE)
    return keyboard

# ===== ПРОВЕРКА ПОДПИСКИ =====

async def check_subscription(user_id: int) -> tuple:
    if not REQUIRED_GROUPS:
        return True, []
    not_subscribed = []
    for group_id in REQUIRED_GROUPS:
        try:
            response = await bot.api.request(
                "groups.isMember",
                {"group_id": group_id, "user_id": user_id, "v": "5.131"}
            )
            if response.get("response") != 1:
                not_subscribed.append(group_id)
        except Exception as e:
            logging.error(f"Ошибка проверки подписки на группу {group_id}: {e}")
            not_subscribed.append(group_id)
    return len(not_subscribed) == 0, not_subscribed

def get_groups_links(groups_ids):
    return "\n".join([f"👉 https://vk.com/club{g}" for g in groups_ids])

# ===== ОБЫЧНЫЕ КОМАНДЫ =====

@bot.on.private_message(text=["начать", "старт", "start", "/start"])
async def start_command(message: Message):
    is_subscribed, not_subscribed = await check_subscription(message.from_id)
    if not is_subscribed:
        await message.answer(
            NOT_SUBSCRIBED_TEXT + get_groups_links(not_subscribed) +
            "\n\nПодпишитесь и нажмите '✅ Проверить подписку'",
            keyboard=get_subscription_keyboard()
        )
        return
    await message.answer(
        "👋 Привет! Я бот для сбора данных.\n\nВыберите действие:",
        keyboard=get_main_keyboard()
    )

@bot.on.private_message(text="✅ Проверить подписку")
async def check_subscription_button(message: Message):
    is_subscribed, not_subscribed = await check_subscription(message.from_id)
    if is_subscribed:
        await message.answer(
            "✅ Отлично! Вы подписаны на все группы.\n\nТеперь вам доступны все функции:",
            keyboard=get_main_keyboard()
        )
    else:
        await message.answer(
            NOT_SUBSCRIBED_TEXT + get_groups_links(not_subscribed) +
            "\n\nПодпишитесь и нажмите кнопку снова",
            keyboard=get_subscription_keyboard()
        )

@bot.on.private_message(text="📝 Регистрация")
async def register_start(message: Message):
    is_subscribed, not_subscribed = await check_subscription(message.from_id)
    if not is_subscribed:
        await message.answer(
            NOT_SUBSCRIBED_TEXT + get_groups_links(not_subscribed) +
            "\n\nПодпишитесь и нажмите '✅ Проверить подписку'",
            keyboard=get_subscription_keyboard()
        )
        return
    
    # ===== ПРОВЕРКА НА ПОВТОРНУЮ РЕГИСТРАЦИЮ =====
    if is_user_registered(message.from_id):
        await message.answer(
            "⚠️ Вы уже зарегистрированы!\n\n"
            "Чтобы посмотреть свои данные, нажмите '📊 Посмотреть данные'.\n"
            "Если хотите изменить данные, обратитесь к администратору.",
            keyboard=get_main_keyboard()
        )
        return
    
    if message.peer_id in user_data_temp:
        del user_data_temp[message.peer_id]
    
    await message.answer(
        "Пожалуйста, введите ваше **ФИО** (полностью):",
        keyboard=get_cancel_keyboard()
    )
    user_data_temp[message.peer_id] = {"state": "waiting_for_full_name"}

@bot.on.private_message(text="📊 Посмотреть данные")
async def view_data(message: Message):
    is_subscribed, not_subscribed = await check_subscription(message.from_id)
    if not is_subscribed:
        await message.answer(
            NOT_SUBSCRIBED_TEXT + get_groups_links(not_subscribed) +
            "\n\nПодпишитесь и нажмите '✅ Проверить подписку'",
            keyboard=get_subscription_keyboard()
        )
        return
    if not os.path.exists(DATA_FILE):
        await message.answer("📭 Данных пока нет.")
        return
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    if ws.max_row == 1:
        await message.answer("📭 Данных пока нет.")
        return
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == message.from_id:
            await message.answer(
                f"📋 Ваши данные:\n\n"
                f"👤 ФИО: {ws.cell(row=row, column=2).value}\n"
                f"📱 Телефон: {ws.cell(row=row, column=3).value}\n"
                f"🎂 Возраст: {ws.cell(row=row, column=4).value}\n"
                f"🏟️ Мероприятие: {ws.cell(row=row, column=5).value}",
                keyboard=get_main_keyboard()
            )
            return
    await message.answer(
        "❌ Вы ещё не зарегистрированы!\nНажмите '📝 Регистрация'",
        keyboard=get_main_keyboard()
    )

@bot.on.private_message(text="❓ Помощь")
async def help_command(message: Message):
    await message.answer(
        "🤖 Инструкция:\n\n"
        "📝 Регистрация - заполнить анкету (доступна 1 раз)\n"
        "📊 Посмотреть данные - показать вашу анкету\n"
        "❌ Отмена - отменить текущее действие\n\n"
        "⚠️ Для использования бота нужно быть подписанным на все группы!\n\n"
        "👑 Админ-команды:\n"
        "!список - показать всех пользователей\n"
        "!очистить - удалить все данные",
        keyboard=get_main_keyboard()
    )

@bot.on.private_message(text="❌ Отмена")
async def cancel_action(message: Message):
    if message.peer_id in user_data_temp:
        del user_data_temp[message.peer_id]
    await message.answer("✅ Действие отменено", keyboard=get_main_keyboard())

# ===== АДМИН-КОМАНДЫ =====

@bot.on.private_message(text="!список")
async def admin_list_users(message: Message):
    """Отправляет список зарегистрированных пользователей"""
    # Проверка прав админа
    if message.from_id not in ADMIN_IDS:
        await message.answer("⛔ У вас нет прав для этой команды.")
        return
    
    # Проверяем, существует ли файл
    if not os.path.exists(DATA_FILE):
        await message.answer("📭 Файл с данными не найден. Зарегистрированных пользователей пока нет.")
        return
    
    # Загружаем файл
    try:
        wb = load_workbook(DATA_FILE)
        ws = wb.active
    except Exception as e:
        await message.answer(f"❌ Ошибка чтения файла: {e}")
        return
    
    # Проверяем, есть ли записи (строки кроме заголовка)
    if ws.max_row <= 1:
        await message.answer("📭 Зарегистрированных пользователей пока нет.")
        return
    
    # Считаем количество пользователей
    total_users = ws.max_row - 1
    
    # Создаём временный файл
    import tempfile
    temp_file = tempfile.NamedTemporaryFile(suffix='.txt', mode='w', encoding='utf-8', delete=False)
    
    try:
        # Записываем заголовок
        temp_file.write("=" * 60 + "\n")
        temp_file.write("     СПИСОК ЗАРЕГИСТРИРОВАННЫХ ПОЛЬЗОВАТЕЛЕЙ\n")
        temp_file.write("=" * 60 + "\n\n")
        
        # Записываем всех пользователей
        for row in range(2, ws.max_row + 1):
            full_name = ws.cell(row=row, column=2).value or "Не указано"
            phone = ws.cell(row=row, column=3).value or "Не указан"
            age = ws.cell(row=row, column=4).value or "Не указан"
            event = ws.cell(row=row, column=5).value or "Не выбрано"
            date = ws.cell(row=row, column=6).value or "Не указана"
            
            temp_file.write(f"#{row-1}\n")
            temp_file.write(f"👤 ФИО: {full_name}\n")
            temp_file.write(f"📱 Телефон: {phone}\n")
            temp_file.write(f"🎂 Возраст: {age} лет\n")
            temp_file.write(f"🏟️ Мероприятие: {event}\n")
            temp_file.write(f"📅 Дата регистрации: {date}\n")
            temp_file.write("─" * 40 + "\n\n")
        
        temp_file.write("\n" + "=" * 60 + "\n")
        temp_file.write(f"  ИТОГО: {total_users} пользователей\n")
        temp_file.write("=" * 60 + "\n")
        temp_file.close()
        
        # Проверяем размер файла
        file_size = os.path.getsize(temp_file.name)
        if file_size == 0:
            await message.answer("❌ Ошибка: создан пустой файл.")
            return
        
        # Отправляем файл
        try:
            from vkbottle.uploader import Uploader
            uploader = Uploader(bot.api)
            
            with open(temp_file.name, 'rb') as f:
                attachment = await uploader.document(
                    document=f,
                    title="users_list.txt",
                    peer_id=message.peer_id
                )
            
            await message.answer(
                f"📊 Всего зарегистрировано: {total_users}\n\n📎 Файл со списком прикреплён ниже.",
                attachment=attachment
            )
            
        except Exception as e:
            # Если файл не отправился - показываем список текстом
            users_text = f"📊 Всего зарегистрировано: {total_users}\n\n"
            users = []
            
            # Показываем до 20 пользователей (чтобы не превысить лимит ВК)
            max_show = min(ws.max_row - 1, 20)
            for row in range(2, max_show + 2):
                full_name = ws.cell(row=row, column=2).value or "Не указано"
                phone = ws.cell(row=row, column=3).value or "Не указан"
                age = ws.cell(row=row, column=4).value or "Не указан"
                event = ws.cell(row=row, column=5).value or "Не выбрано"
                date = ws.cell(row=row, column=6).value or "Не указана"
                
                users.append(
                    f"👤 {full_name}\n"
                    f"📱 {phone}\n"
                    f"🎂 {age} лет\n"
                    f"🏟️ {event}\n"
                    f"📅 {date}\n"
                    f"{'─'*15}"
                )
            
            users_text += "\n\n".join(users)
            
            if total_users > 20:
                users_text += f"\n\n... и ещё {total_users - 20} пользователей."
            
            await message.answer(users_text)
            
    except Exception as e:
        await message.answer(f"❌ Ошибка: {e}")
    
    finally:
        # Удаляем временный файл
        try:
            os.unlink(temp_file.name)
        except:
            pass

@bot.on.private_message(text="!очистить")
async def admin_clear_users(message: Message):
    if message.from_id not in ADMIN_IDS:
        await message.answer("⛔ У вас нет прав для этой команды.")
        return
    if not os.path.exists(DATA_FILE):
        await message.answer("📭 Данных пока нет.")
        return
    await message.answer(
        "⚠️ ВНИМАНИЕ! Вы уверены, что хотите удалить ВСЕ данные?\n\n"
        "Для подтверждения напишите: **да, очистить**\n"
        "Для отмены напишите: **отмена**"
    )
    user_data_temp[message.peer_id] = {"state": "waiting_for_clear"}

# ===== ГЛАВНЫЙ ОБРАБОТЧИК СОСТОЯНИЙ =====

@bot.on.private_message()
async def handle_all_messages(message: Message):
    if message.peer_id not in user_data_temp:
        await message.answer(
            "❓ Неизвестная команда.\nИспользуйте кнопки.",
            keyboard=get_main_keyboard()
        )
        return

    current_state = user_data_temp[message.peer_id].get("state")

    # ===== ПОДТВЕРЖДЕНИЕ ОЧИСТКИ =====
    if current_state == "waiting_for_clear":
        text = message.text.strip().lower()
        if text == "да, очистить":
            if os.path.exists(DATA_FILE):
                os.remove(DATA_FILE)
            init_excel()
            await message.answer("✅ Все данные успешно очищены!")
            del user_data_temp[message.peer_id]
            return
        elif text == "отмена":
            await message.answer("❌ Очистка отменена.")
            del user_data_temp[message.peer_id]
            return
        else:
            await message.answer(
                "❌ Неверная команда.\n"
                "Для подтверждения напишите: **да, очистить**\n"
                "Для отмены напишите: **отмена**"
            )
            return

    # ===== ФИО =====
    if current_state == "waiting_for_full_name":
        full_name = message.text.strip()
        if len(full_name.split()) < 2:
            await message.answer("❌ Пожалуйста, введите полное ФИО (Фамилия Имя Отчество):")
            return
        user_data_temp[message.peer_id]["full_name"] = full_name
        user_data_temp[message.peer_id]["state"] = "waiting_for_phone"
        await message.answer(
            "📱 Введите ваш номер телефона:",
            keyboard=get_cancel_keyboard()
        )

    # ===== ТЕЛЕФОН =====
    elif current_state == "waiting_for_phone":
        phone = message.text.strip()
        if len(phone) < 10:
            await message.answer("❌ Некорректный номер. Попробуйте ещё раз:")
            return
        user_data_temp[message.peer_id]["phone"] = phone
        user_data_temp[message.peer_id]["state"] = "waiting_for_age"
        await message.answer(
            "🎂 Введите ваш возраст:",
            keyboard=get_cancel_keyboard()
        )

    # ===== ВОЗРАСТ =====
    elif current_state == "waiting_for_age":
        try:
            age = int(message.text.strip())
            if age < 1 or age > 150:
                raise ValueError
        except ValueError:
            await message.answer("❌ Введите число от 1 до 150:")
            return
        user_data_temp[message.peer_id]["age"] = age
        user_data_temp[message.peer_id]["state"] = "waiting_for_event"
        await message.answer(
            "🏟️ Выберите мероприятие:",
            keyboard=get_event_keyboard()
        )

    # ===== МЕРОПРИЯТИЕ =====
    elif current_state == "waiting_for_event":
        event = message.text.strip()
        if event not in EVENTS and event != "❌ Отмена":
            await message.answer(
                "❌ Пожалуйста, выберите мероприятие из списка кнопок:",
                keyboard=get_event_keyboard()
            )
            return
        
        if event == "❌ Отмена":
            del user_data_temp[message.peer_id]
            await message.answer("❌ Регистрация отменена.", keyboard=get_main_keyboard())
            return
        
        data = user_data_temp[message.peer_id]
        save_user_data(
            message.from_id,
            data.get("full_name", "Не указано"),
            data.get("phone", "Не указано"),
            data.get("age", "Не указано"),
            event
        )
        del user_data_temp[message.peer_id]
        await message.answer(
            f"✅ Регистрация на мероприятие **{event}** завершена! 🎉\n\n"
            f"Ваши данные:\n"
            f"👤 ФИО: {data.get('full_name')}\n"
            f"📱 Телефон: {data.get('phone')}\n"
            f"🎂 Возраст: {data.get('age')} лет\n"
            f"🏟️ Мероприятие: {event}\n\n"
            f"Теперь вы можете посмотреть свои данные через '📊 Посмотреть данные'",
            keyboard=get_main_keyboard()
        )

# ===== ЗАПУСК =====

if __name__ == "__main__":
    init_excel()
    print("🤖 Бот запущен!")
    print(f"✅ Проверяются группы: {REQUIRED_GROUPS}")
    print(f"✅ Администраторы: {ADMIN_IDS}")
    print("✅ Нажмите Ctrl+C для остановки")
    asyncio.run(bot.run_polling())
